"""
Moduł debugowy scrapera RemoteOK.
Wywołuje scrape_remoteok() i zapisuje wyniki do XLSX w ./debug/.

Uruchomienie: py -3.12 debug_remoteok.py
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

DEBUG_DIR = Path("debug")

_COLS = [
    ("title",        35),
    ("company",      22),
    ("location",     18),
    ("salary",       25),
    ("skills",       40),
    ("url",          55),
    ("desc_len",     10),
    ("description",  80),
]
_CENTER = {7}


def _save_xlsx(offers: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RemoteOK"

    hdr_fill = PatternFill("solid", fgColor="37474F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    wrap     = Alignment(vertical="top", wrap_text=True)
    center   = Alignment(horizontal="center", vertical="top")
    url_font = Font(color="1A73E8", underline="single", size=9)

    for ci, (header, width) in enumerate(_COLS, 1):
        cell           = ws.cell(row=1, column=ci, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = "A2"

    for ri, o in enumerate(offers, 2):
        desc = o.get("description") or ""
        row_data = [
            o.get("title", ""),
            o.get("company", ""),
            o.get("location", ""),
            o.get("salary", ""),
            ", ".join(o.get("skills") or []),
            o.get("url", ""),
            len(desc),
            desc[:500],
        ]
        for ci, val in enumerate(row_data, 1):
            cell           = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = center if ci in _CENTER else wrap

        url_val = o.get("url", "")
        if url_val:
            c           = ws.cell(row=ri, column=6)
            c.hyperlink = url_val
            c.font      = url_font

        ws.row_dimensions[ri].height = 45 if desc else 20

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}1"
    wb.save(path)
    logger.info(f"[XLSX] Zapisano: {path}  ({len(offers)} wierszy)")


def run() -> None:
    from modules.scraper import scrape_remoteok

    DEBUG_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    logger.info("=== DEBUG REMOTEOK SCRAPER START ===")
    offers = scrape_remoteok()

    n_desc = sum(1 for o in offers if o.get("description"))
    n_sal  = sum(1 for o in offers if o.get("salary"))
    avg_d  = (sum(len(o.get("description") or "") for o in offers if o.get("description"))
              // max(n_desc, 1))

    logger.info("=" * 55)
    logger.info(f"WYNIKI  –  {len(offers)} ofert łącznie")
    logger.info(f"  z opisem       : {n_desc}  ({n_desc*100//max(len(offers),1)}%)")
    logger.info(f"  z salary (USD) : {n_sal}  ({n_sal*100//max(len(offers),1)}%)")
    logger.info(f"  avg opis [zn.] : {avg_d}")
    logger.info("=" * 55)

    print(f"\n{'#':>3}  {'Tytuł':<38} {'Firma':<22} {'Salary':<22} {'Opis'}")
    print("-" * 110)
    for i, o in enumerate(offers, 1):
        d = len(o.get("description") or "")
        print(f"{i:>3}  {'📄' if d else '  '} {o.get('title','')[:36]:<38} "
              f"{o.get('company','')[:20]:<22} "
              f"{(o.get('salary') or '—')[:20]:<22} "
              f"{d}zn")

    xlsx_path = DEBUG_DIR / f"remoteok_{ts}.xlsx"
    _save_xlsx(offers, xlsx_path)
    logger.info(f"XLSX → {xlsx_path.resolve()}")
    logger.info("=== DEBUG REMOTEOK SCRAPER KONIEC ===")


if __name__ == "__main__":
    run()
