"""
Job Hunter – tryb lokalny
Uruchomienie: python main.py
Flagi:
  --no-analyze   tylko scraping, bez analizy Claude (szybszy test)
  --limit N      ogranicz do N ofert przed analizą
"""

import sys
import json
import logging
import argparse
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
import openpyxl
from modules.company_urls import get_company_url
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Force UTF-8 output on Windows terminals
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

load_dotenv()

_LOGS_DIR = Path("logs")
_LOGS_DIR.mkdir(exist_ok=True)
_log_file = _LOGS_DIR / f"job_hunter_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(_log_file, encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

RESULTS_DIR = Path("results")

# Kolory ANSI dla terminala
GREEN  = "\033[92m"
YELLOW = "\033[93m"
RED    = "\033[91m"
CYAN   = "\033[96m"
BOLD   = "\033[1m"
RESET  = "\033[0m"
DIM    = "\033[2m"


def _score_color(score) -> str:
    if score >= 8:
        return GREEN
    if score >= 6:
        return YELLOW
    return DIM


def _display_score(o: dict) -> str:
    """Zwraca final_score (float) jeśli dostępny, inaczej score (int)."""
    fs = o.get("final_score")
    if fs is not None:
        return f"{fs:.2f}"
    s = o.get("score")
    return str(s) if s is not None else "?"


def _print_offer(idx: int, o: dict, dim: bool = False) -> None:
    score_val = o.get("final_score") if o.get("final_score") is not None else o.get("score", 0)
    score_str = _display_score(o)
    verdict   = o.get("verdict", "?")
    sc        = _score_color(score_val) if not dim else DIM
    predicted = " [est.]" if o.get("salary_predicted") else ""
    emphasis  = o.get("cv_emphasis", "—")
    source    = o.get("source", "—")
    prefix    = DIM if dim else ""
    reset_    = RESET

    print(f"{prefix}{BOLD}#{idx:02d}  {sc}{score_str}/10{reset_}{prefix}  [{verdict}]  {BOLD if not dim else ''}{o.get('title', '')}{reset_}")
    print(f"{prefix}      Firma:    {o.get('company', '—')}{reset_}")
    print(f"{prefix}      Lokacja:  {o.get('location', '—')}{reset_}")
    print(f"{prefix}      Zarobki:  {o.get('salary', '—')}{predicted}{reset_}")
    print(f"{prefix}      Tryb CV:  {emphasis}  |  Źródło: {source}{reset_}")
    print(f"{prefix}      Powód:    {o.get('match_reason', '')}{reset_}")
    print(f"{prefix}      URL:      {'' if dim else CYAN}{o.get('url', '')}{reset_}")
    print()


def print_results(offers: list[dict]) -> None:
    apply_list = [o for o in offers if o.get("verdict") == "APPLY"]
    skip_list  = [o for o in offers if o.get("verdict") != "APPLY"]

    print(f"\n{BOLD}{'═' * 90}{RESET}")
    print(f"{BOLD}{CYAN}  JOB HUNTER – wyniki  "
          f"({len(apply_list)} APPLY  /  {len(skip_list)} SKIP  /  {len(offers)} łącznie){RESET}")
    print(f"{BOLD}{'═' * 90}{RESET}\n")

    for i, o in enumerate(apply_list, 1):
        _print_offer(i, o, dim=False)

    if skip_list:
        print(f"{DIM}{'─' * 90}")
        print(f"  SKIP – score < 6  ({len(skip_list)} ofert poniżej progu, spełniły warunki lokalizacji/wynagrodzenia)")
        print(f"{'─' * 90}{RESET}\n")
        for i, o in enumerate(skip_list, 1):
            _print_offer(i, o, dim=True)

    print(f"{DIM}Legenda: [est.] = przewidywane wynagrodzenie{RESET}\n")


def save_results(offers: list[dict]) -> tuple[Path, Path]:
    RESULTS_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    json_path = RESULTS_DIR / f"offers_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(offers, f, ensure_ascii=False, indent=2)

    xls_path = RESULTS_DIR / f"offers_{timestamp}.xlsx"
    _save_xlsx(offers, xls_path)

    return json_path, xls_path


def _salary_type(o: dict) -> str:
    if not o.get("salary"):
        return "brak"
    if not o.get("salary_predicted"):
        return "ogłoszenie"
    contract = (o.get("salary_contract") or "").upper()
    model = "claude" if "est.reguły" not in o.get("salary", "").lower() else "reguły"
    return f"AI est. ({model}{', ' + contract if contract else ''})"


def _save_xlsx(offers: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oferty"

    # ── Style ──────────────────────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1A73E8")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    apply_fill = PatternFill("solid", fgColor="E8F5E9")
    skip_fill  = PatternFill("solid", fgColor="F5F5F5")
    url_font   = Font(color="1A73E8", underline="single", size=9)
    center     = Alignment(horizontal="center", vertical="top", wrap_text=False)
    wrap       = Alignment(vertical="top", wrap_text=True)
    thin       = Side(style="thin", color="DDDDDD")
    border     = Border(bottom=thin)

    # ── Nagłówki ───────────────────────────────────────────────────────────────
    COLS = [
        ("Score",                    8),
        ("Verdict",                  8),
        ("Stanowisko",              28),
        ("Firma",                   22),
        ("Lokalizacja",             18),
        ("Wynagrodzenie PLN/mies.", 28),
        ("Typ wynagrodzenia",       16),
        ("Źródło",                  13),
        ("Akcent CV",               12),
        ("Powód dopasowania",       42),
        ("Luki kandydata",          40),
        ("Opis oferty",             70),
        ("URL oferty",              50),
        ("Strona kariery firmy",    40),
    ]
    for col_idx, (header, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    # ── Dane ───────────────────────────────────────────────────────────────────
    for row_idx, o in enumerate(offers, 2):
        is_apply    = o.get("verdict") == "APPLY"
        fill        = apply_fill if is_apply else skip_fill
        score       = o.get("final_score") if o.get("final_score") is not None else o.get("score", "")
        offer_url   = o.get("url", "")
        company     = o.get("company", "")
        company_url = get_company_url(company)

        desc_cell = (o.get("description") or "")[:500]
        key_gaps  = "; ".join(o.get("key_gaps") or [])

        values = [
            score,
            o.get("verdict", ""),
            o.get("title", ""),
            company,
            o.get("location", ""),
            o.get("salary", ""),
            _salary_type(o),
            o.get("source", ""),
            o.get("cv_emphasis", ""),
            o.get("match_reason", ""),
            key_gaps,
            desc_cell,
            offer_url,
            company_url,
        ]

        for col_idx, value in enumerate(values, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col_idx in (1, 2, 8, 9) else wrap

        # Score – pogrubiony + kolor tekstu (działa dla int i float)
        if isinstance(score, (int, float)):
            ws.cell(row=row_idx, column=1).font = Font(
                bold=True,
                color=("00AA44" if score >= 8 else "FFA000" if score >= 6 else "888888"),
            )

        # URL oferty – klikalne hiperłącze (kolumna 13 po dodaniu Luki)
        if offer_url:
            c = ws.cell(row=row_idx, column=13)
            c.hyperlink = offer_url
            c.value     = offer_url
            c.font      = url_font

        # Strona kariery – klikalne hiperłącze (kolumna 14)
        if company_url:
            c = ws.cell(row=row_idx, column=14)
            c.hyperlink = company_url
            c.value     = company_url
            c.font      = url_font

        ws.row_dimensions[row_idx].height = 60 if desc_cell else 28

    # ── Autofiltr ──────────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    wb.save(path)


def _check_env():
    from modules.analyzer import claude_available
    key = claude_available()
    if not key:
        print(f"\n{YELLOW}[!] CLAUDE_API_KEY nie ustawiony lub nieprawidłowy{RESET}")
        print(f"{DIM}   Scoring i predykcja płac działają w trybie REGUŁ (bez AI).{RESET}")
        print(f"{DIM}   Lokalnie: ustaw CLAUDE_API_KEY=sk-ant-... w pliku .env{RESET}")
        print(f"{DIM}   Cloud Run: zaktualizuj sekret CLAUDE_API_KEY w Secret Manager\n{RESET}")
    else:
        print(f"\n{GREEN}[OK] Claude API aktywny – pełna analiza AI{RESET}")


def _db_available() -> bool:
    import os
    return all(os.environ.get(k) for k in ("DB_HOST", "DB_NAME", "DB_USER", "DB_PASSWORD"))


def run(no_analyze: bool = False, limit: int = None,
        portals: list[str] | None = None, ignore_seen: bool = False):
    from modules.scraper import scrape_all
    from modules.analyzer import analyze_all, _normalize_salary_pln

    _check_env()
    logger.info("=== Job Hunter START ===")

    # DB – opcjonalne; lokalnie bez Cloud SQL po prostu pomijamy
    db_ok = _db_available()
    if db_ok:
        from modules.db import init_db, get_known_urls, insert_offers, mark_notified, update_offer_status
        try:
            init_db()
        except Exception as e:
            logger.warning(f"[DB] init_db błąd: {e} – kontynuuję bez DB")
            db_ok = False
    else:
        logger.warning("[DB] Brak kredencjałów DB – kroki DB pominięte")

    # 1. Scraping
    portal_info = f" (portal: {portals})" if portals else ""
    logger.info(f"Krok 1: Scraping ofert{portal_info}...")
    offers = scrape_all(portals=portals)

    if not offers:
        print(f"\n{RED}Brak ofert po scrapingu.{RESET}")
        return

    total_scraped = len(offers)
    print(f"\n{GREEN}Pobrano {total_scraped} unikalnych ofert.{RESET}")

    if db_ok:
        try:
            update_offer_status({o.get("url") for o in offers if o.get("url")})
        except Exception as e:
            logger.warning(f"[DB] update_offer_status błąd: {e}")

    if limit:
        offers = offers[:limit]
        logger.info(f"Ograniczono do {limit} ofert (--limit)")

    # Normalizacja walut zawsze – niezależnie od trybu
    offers = [_normalize_salary_pln(o) for o in offers]

    if ignore_seen:
        new_offers = offers
        logger.info(f"--ignore-seen: analiza wszystkich {len(offers)} ofert (bez filtra known_urls)")
    elif db_ok:
        try:
            known_urls = get_known_urls()
            new_offers = [o for o in offers if o.get("url") not in known_urls]
            logger.info(f"Nowe oferty: {len(new_offers)} / {len(offers)} łącznie")
        except Exception as e:
            logger.warning(f"[DB] get_known_urls błąd: {e} – traktuję wszystkie jako nowe")
            new_offers = offers
    else:
        new_offers = offers
        logger.info(f"Brak DB – traktuję wszystkie {len(offers)} ofert jako nowe")

    if not new_offers:
        logger.info("Brak nowych ofert – kończę")
        return

    if no_analyze:
        print(f"{YELLOW}Tryb --no-analyze: brak scoringu i predykcji płac.{RESET}")
        print_results(new_offers)
        json_path, xls_path = save_results(new_offers)
        print(f"Zapisano: {json_path}")
        print(f"Zapisano: {xls_path}")
        return

    # 2. Analiza (Claude lub reguły)
    logger.info(f"Krok 2: Analiza {len(new_offers)} ofert...")
    scored = analyze_all(new_offers)

    if not scored:
        print(f"\n{YELLOW}Brak ofert po filtrze wynagrodzenia.{RESET}")
        print(f"{DIM}Spróbuj: python main.py --no-analyze  żeby zobaczyć surowe wyniki.{RESET}")
        return

    apply_count = sum(1 for o in scored if o.get("verdict") == "APPLY")
    print(f"\n{GREEN}Wyniki: {apply_count} APPLY  /  {len(scored) - apply_count} SKIP  "
          f"/  {len(scored)} łącznie po filtrze wynagrodzenia{RESET}")

    # 3. Wyniki
    if db_ok:
        try:
            insert_offers(scored)
        except Exception as e:
            logger.warning(f"[DB] insert_offers błąd: {e}")
    print_results(scored)
    json_path, xls_path = save_results(scored)
    print(f"{DIM}Wyniki zapisano do: {json_path}{RESET}")
    print(f"{DIM}              XLSX: {xls_path}{RESET}")

    # 4. E-mail z raportem
    from modules.emailer import send_report, _email_available
    if _email_available():
        send_report(scored, xls_path, total_scraped=total_scraped)
        print(f"{GREEN}[OK] Raport wysłany e-mailem{RESET}")
    else:
        print(f"{YELLOW}[!] E-mail pominięty – ustaw GMAIL_APP_PASSWORD w .env{RESET}")

    if db_ok:
        try:
            to_notify = [o for o in scored if o.get("verdict") == "APPLY"]
            mark_notified(to_notify)
        except Exception as e:
            logger.warning(f"[DB] mark_notified błąd: {e}")

    logger.info(f"=== Job Hunter KONIEC – {apply_count} APPLY / {len(scored)} łącznie ===")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Job Hunter – lokalny scraper ofert")
    parser.add_argument("--no-analyze",   action="store_true", help="Pomiń analizę Claude (tylko scraping)")
    parser.add_argument("--limit",        type=int, default=None, help="Ogranicz liczbę ofert do analizy")
    parser.add_argument("--portal",       type=str, nargs="+", default=None,
                        help="Uruchom tylko wskazane portale, np. --portal jjit nofluff")
    parser.add_argument("--ignore-seen",  action="store_true",
                        help="Analizuj wszystkie oferty, nawet te już w DB (do testów)")
    args = parser.parse_args()
    run(
        no_analyze=args.no_analyze,
        limit=args.limit,
        portals=args.portal,  # już lista dzięki nargs="+"
        ignore_seen=args.ignore_seen,
    )
