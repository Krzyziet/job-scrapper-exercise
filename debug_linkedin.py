"""
Moduł debugowy scrapera LinkedIn.
Wywołuje scrape_linkedin() i zapisuje próbkę 10 ofert do XLSX w ./debug/.

Uruchomienie: py -3.12 debug_linkedin.py
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

DEBUG_DIR = Path("debug")

_COLS = [
    ("title",        38),
    ("company",      22),
    ("location",     20),
    ("salary",       15),
    ("skills",       30),
    ("url",          55),
    ("desc_len",     10),
    ("description",  80),
]
_CENTER = {7}


def _save_xlsx(offers: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LinkedIn"

    hdr_fill = PatternFill("solid", fgColor="0A66C2")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    wrap     = Alignment(vertical="top", wrap_text=True)
    center   = Alignment(horizontal="center", vertical="top")
    url_font = Font(color="1A73E8", underline="single", size=9)

    for ci, (header, width) in enumerate(_COLS, 1):
        cell           = ws.cell(row=1, column=ci, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = "A2"

    for ri, o in enumerate(offers, 2):
        desc = o.get("description") or ""
        row_data = [
            o.get("title", ""),
            o.get("company", ""),
            o.get("location", ""),
            o.get("salary", "") or "—",
            ", ".join(o.get("skills") or []),
            o.get("url", ""),
            len(desc),
            desc[:500],
        ]
        for ci, val in enumerate(row_data, 1):
            cell           = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = center if ci in _CENTER else wrap

        url_val = o.get("url", "")
        if url_val:
            c           = ws.cell(row=ri, column=6)
            c.hyperlink = url_val
            c.font      = url_font

        ws.row_dimensions[ri].height = 40 if desc else 20

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}1"
    wb.save(path)
    logger.info(f"[XLSX] Zapisano: {path}  ({len(offers)} wierszy)")


def run() -> None:
    from modules.scraper import scrape_linkedin

    DEBUG_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    logger.info("=== DEBUG LinkedIn SCRAPER START ===")
    offers = scrape_linkedin()

    sample = offers[:10]

    n_desc = sum(1 for o in sample if o.get("description"))
    n_sal  = sum(1 for o in sample if o.get("salary"))

    logger.info("=" * 55)
    logger.info(f"Łącznie scraped : {len(offers)} ofert")
    logger.info(f"Próbka do XLSX  : {len(sample)} ofert")
    logger.info(f"Z opisem        : {n_desc}")
    logger.info(f"Z salary        : {n_sal}")
    logger.info("=" * 55)

    print(f"\n{'#':>3}  {'Tytuł':<38} {'Firma':<22} {'Lokacja':<20}")
    print("-" * 90)
    for i, o in enumerate(sample, 1):
        print(f"{i:>3}  {o.get('title','')[:36]:<38} "
              f"{o.get('company','')[:20]:<22} "
              f"{o.get('location','')[:18]:<20}")

    xlsx_path = DEBUG_DIR / f"linkedin_{ts}.xlsx"
    _save_xlsx(sample, xlsx_path)
    logger.info(f"XLSX → {xlsx_path.resolve()}")
    logger.info("=== DEBUG LinkedIn KONIEC ===")


if __name__ == "__main__":
    run()
