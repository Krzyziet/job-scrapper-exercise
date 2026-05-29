"""
Narzędzie diagnostyczne – JustJoinIT scraping bez DB, bez AI.
Pobiera WSZYSTKIE oferty przez REST API (paginacja cursor-based).

Uruchomienie: python scrape_debug.py
"""

import json
import logging
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

from modules.scraper_pw import _make_browser, _matches_role, _matches_location, _salary_ok

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

DEBUG_DIR    = Path("debug")
SEARCH_TERMS = ["product owner", "chapter lead", "engineering manager"]
PAGE_SIZE    = 20    # JJIT zwraca ~10 niezależnie od tej wartości
API_DELAY    = 0.3   # sekund między requestami


# ── Salary ────────────────────────────────────────────────────────────────────

def _parse_employment_types(et_list: list) -> tuple[int, int, str, str]:
    """
    Zwraca (salary_from, salary_to, currency, contract_type).
    Bierze tylko currencySource='original'; priorytet: b2b > permanent > reszta.
    """
    if not et_list:
        return 0, 0, "", ""

    original = [et for et in et_list if et.get("currencySource") == "original"] or et_list
    _PRIO = {"b2b": 0, "b2b_contract": 0, "permanent": 1,
             "contract_of_employment": 1, "mandate_contract": 2}
    ordered = sorted(original, key=lambda x: _PRIO.get(x.get("type", ""), 9))

    for et in ordered:
        sf = et.get("from") or 0
        st = et.get("to")   or 0
        if sf:
            return int(sf), int(st), et.get("currency", "") or "", et.get("type", "") or ""

    return 0, 0, "", (ordered[0].get("type", "") if ordered else "")


# ── API ───────────────────────────────────────────────────────────────────────

def _fetch_page(page, term: str, cursor: int | None) -> dict | None:
    url = (
        f"https://justjoin.it/api/candidate-api/offers"
        f"?keywords={term.replace(' ', '+')}&keywordType=any&pageSize={PAGE_SIZE}"
        + (f"&from={cursor}" if cursor is not None else "")
    )
    try:
        resp = page.goto(url, timeout=20_000, wait_until="domcontentloaded")
        if resp and resp.status == 200:
            return resp.json()
        logger.warning(f"[API] HTTP {resp.status if resp else '?'}  {url}")
    except Exception as e:
        logger.warning(f"[API] błąd: {e}  {url}")
    return None


def _offer_from_item(item: dict, search_term: str) -> dict:
    slug    = item.get("slug") or item.get("guid") or ""
    et_list = item.get("employmentTypes") or []
    sf, st, cur, ctype = _parse_employment_types(et_list)
    city = item.get("city") or ""
    wp   = item.get("workplaceType") or ""
    return {
        "search_term":    search_term,
        "title":          item.get("title") or "",
        "company":        item.get("companyName") or "",
        "city":           city,
        "workplace_type": wp,
        "url":            f"https://justjoin.it/job-offer/{slug}" if slug else "",
        "salary_from":    sf,
        "salary_to":      st,
        "currency":       cur,
        "contract_type":  ctype,
        "salary_raw":     json.dumps(et_list, ensure_ascii=False),
        "raw_keys":       json.dumps(sorted(item.keys())),
    }


# ── Główna funkcja ────────────────────────────────────────────────────────────

def debug_justjoinit() -> None:
    DEBUG_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    all_offers: list[dict] = []
    seen_urls:  set[str]   = set()
    api_file_n = 0

    with sync_playwright() as pw:
        browser, ctx = _make_browser(pw)
        page = ctx.new_page()

        for term in SEARCH_TERMS:
            cursor   = None
            pg_num   = 0
            term_n   = 0

            while True:
                pg_num += 1
                raw = _fetch_page(page, term, cursor)
                if not raw:
                    break

                offers_raw = raw.get("data", [])
                meta       = raw.get("meta", {})
                total_int  = meta.get("totalItems") or 0

                logger.info(f"[API] '{term}'  s.{pg_num}  +{len(offers_raw)} ofert  "
                            f"(łącznie na JJIT: {total_int})")

                api_file_n += 1
                new_this_page = 0
                for item in offers_raw:
                    if not isinstance(item, dict):
                        continue
                    o = _offer_from_item(item, term)
                    if o["url"] and o["url"] not in seen_urls:
                        seen_urls.add(o["url"])
                        all_offers.append(o)
                        term_n += 1
                        new_this_page += 1

                cursor = (meta.get("next") or {}).get("cursor")

                # Warunek stopu: brak nowych ofert LUB cursor poza zakresem LUB brak cursora
                if (new_this_page == 0
                        or cursor is None
                        or (total_int and cursor >= total_int)):
                    break
                time.sleep(API_DELAY)

            logger.info(f"[API] '{term}' → {term_n} unikalnych ofert")

        browser.close()

    # Kolumny filtrów (boolean, bez odrzucania)
    for o in all_offers:
        loc_check = o.get("workplace_type", "") or o.get("city", "")
        o["would_pass_role"]     = _matches_role(o["title"])
        o["would_pass_location"] = _matches_location(loc_check)
        o["would_pass_salary"]   = _salary_ok(o["salary_from"], o["contract_type"])

    # ── XLSX ──────────────────────────────────────────────────────────────────
    xlsx_path = DEBUG_DIR / f"jjit_raw_{timestamp}.xlsx"
    _save_xlsx(all_offers, xlsx_path)

    # ── Podsumowanie ──────────────────────────────────────────────────────────
    n_sal  = sum(1 for o in all_offers if o["salary_from"] > 0)
    n_role = sum(1 for o in all_offers if o["would_pass_role"])
    n_loc  = sum(1 for o in all_offers if o["would_pass_location"])
    n_sf   = sum(1 for o in all_offers if o["would_pass_salary"])
    pct    = n_sal * 100 // max(len(all_offers), 1)

    logger.info("=" * 60)
    logger.info(f"PODSUMOWANIE  –  {len(all_offers)} unikalnych ofert")
    logger.info(f"  salary_from > 0       : {n_sal}  ({pct}%)")
    logger.info(f"  would_pass_role       : {n_role}")
    logger.info(f"  would_pass_location   : {n_loc}")
    logger.info(f"  would_pass_salary     : {n_sf}")
    logger.info(f"XLSX  →  {xlsx_path.resolve()}")
    logger.info(f"Stron API łącznie  →  {api_file_n}")
    logger.info("=" * 60)


# ── XLSX ──────────────────────────────────────────────────────────────────────

_COLS = [
    ("search_term",         16),
    ("title",               32),
    ("company",             22),
    ("city",                16),
    ("workplace_type",      14),
    ("url",                 55),
    ("salary_from",         12),
    ("salary_to",           12),
    ("currency",            10),
    ("contract_type",       16),
    ("salary_raw",          55),
    ("raw_keys",            65),
    ("would_pass_role",     14),
    ("would_pass_location", 16),
    ("would_pass_salary",   16),
]

_CENTER_COLS = {1, 7, 8, 9, 10, 13, 14, 15}


def _save_xlsx(offers: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JJIT raw"

    hdr_fill = PatternFill("solid", fgColor="1A73E8")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    wrap     = Alignment(vertical="top", wrap_text=True)
    center   = Alignment(horizontal="center", vertical="top")

    for ci, (header, width) in enumerate(_COLS, 1):
        cell           = ws.cell(row=1, column=ci, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = "A2"

    for ri, o in enumerate(offers, 2):
        for ci, (key, _) in enumerate(_COLS, 1):
            val            = o.get(key, "")
            if isinstance(val, bool):
                val = "TAK" if val else "nie"
            cell           = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = center if ci in _CENTER_COLS else wrap
        ws.row_dimensions[ri].height = 30

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}1"
    wb.save(path)
    logger.info(f"[XLSX] Zapisano: {path}  ({len(offers)} wierszy)")


if __name__ == "__main__":
    debug_justjoinit()
