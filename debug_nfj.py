"""
Moduł debugowy scrapera NoFluffJobs.
Wywołuje scrape_nofluffjobs() i zapisuje wyniki do XLSX w ./debug/.

Uruchomienie: py -3.12 debug_nfj.py
NIE modyfikuje main.py ani bazy danych.
"""

import logging
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

DEBUG_DIR = Path("debug")

_COLS = [
    ("title",           32),
    ("company",         22),
    ("location",        22),
    ("salary",          32),
    ("salary_from",     12),
    ("salary_contract", 14),
    ("skills",          35),
    ("url",             55),
    ("desc_len",        10),
    ("description",     80),
]
_CENTER = {5, 6, 9}


def _save_xlsx(offers: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NFJ scraper"

    hdr_fill = PatternFill("solid", fgColor="E65100")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    wrap     = Alignment(vertical="top", wrap_text=True)
    center   = Alignment(horizontal="center", vertical="top")

    for ci, (header, width) in enumerate(_COLS, 1):
        cell           = ws.cell(row=1, column=ci, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = "A2"

    url_font = Font(color="1A73E8", underline="single", size=9)

    for ri, o in enumerate(offers, 2):
        desc = o.get("description") or ""
        row_data = [
            o.get("title", ""),
            o.get("company", ""),
            o.get("location", ""),
            o.get("salary", ""),
            o.get("salary_from") or 0,
            o.get("salary_contract", ""),
            ", ".join(o.get("skills") or []),
            o.get("url", ""),
            len(desc),
            desc[:500],
        ]
        for ci, val in enumerate(row_data, 1):
            cell           = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = center if ci in _CENTER else wrap

        url_val = o.get("url", "")
        if url_val:
            c           = ws.cell(row=ri, column=8)
            c.hyperlink = url_val
            c.font      = url_font

        ws.row_dimensions[ri].height = 45 if desc else 20

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}1"
    wb.save(path)
    logger.info(f"[XLSX] Zapisano: {path}  ({len(offers)} wierszy)")


def run() -> None:
    from modules.scraper import scrape_nofluffjobs

    DEBUG_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    logger.info("=== DEBUG NFJ SCRAPER START ===")
    offers = scrape_nofluffjobs()

    n_sal  = sum(1 for o in offers if o.get("salary_from", 0) > 0)
    n_desc = sum(1 for o in offers if o.get("description"))
    n_rem  = sum(1 for o in offers if "remote" in (o.get("location") or "").lower())
    avg_d  = (sum(len(o.get("description") or "") for o in offers if o.get("description"))
              // max(n_desc, 1))

    logger.info("=" * 55)
    logger.info(f"WYNIKI  –  {len(offers)} ofert łącznie")
    logger.info(f"  z salary_from > 0 : {n_sal}  ({n_sal*100//max(len(offers),1)}%)")
    logger.info(f"  z opisem          : {n_desc}  ({n_desc*100//max(len(offers),1)}%)")
    logger.info(f"  avg opis [zn.]    : {avg_d}")
    logger.info(f"  remote            : {n_rem}")
    logger.info("=" * 55)

    print(f"\n{'#':>3}  {'Tytuł':<38} {'Firma':<22} {'Lokacja':<20} {'Salary'}")
    print("-" * 120)
    for i, o in enumerate(offers, 1):
        desc_ind = "📄" if o.get("description") else "  "
        print(f"{i:>3}  {desc_ind} {o.get('title','')[:36]:<38} "
              f"{o.get('company','')[:20]:<22} "
              f"{o.get('location','')[:18]:<20} "
              f"{o.get('salary','—')[:35]}")

    xlsx_path = DEBUG_DIR / f"nfj_scraper_{ts}.xlsx"
    _save_xlsx(offers, xlsx_path)
    logger.info(f"XLSX → {xlsx_path.resolve()}")
    logger.info("=== DEBUG NFJ SCRAPER KONIEC ===")


if __name__ == "__main__":
    run()
