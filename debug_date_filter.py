"""
Moduł debugowy filtra dat.
Sprawdza dla każdego scrapera z filtrem dat:
  - ile ofert zwraca API łącznie
  - ile przechodzi filtr (ostatnie 7 dni)
  - rozkład dat publikacji

Uruchomienie: py -3.12 debug_date_filter.py
NIE dotyka main.py ani bazy danych.
"""

import sys
import io
import logging
import time
import requests
import json as _json
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

DEBUG_DIR = Path("debug")

from modules.scraper import _is_recent, SCRAPER_DAYS

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_date(val) -> datetime | None:
    if not val:
        return None
    from email.utils import parsedate_to_datetime
    try:
        if isinstance(val, (int, float)):
            return datetime.fromtimestamp(float(val), tz=timezone.utc)
        s = str(val).strip()
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        try:
            dt = datetime.fromisoformat(s)
        except ValueError:
            dt = parsedate_to_datetime(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None


def _age_label(dt: datetime | None) -> str:
    if dt is None:
        return "brak daty"
    now = datetime.now(timezone.utc)
    days = (now - dt).days
    if days == 0:
        return "dziś"
    if days == 1:
        return "wczoraj"
    if days <= 7:
        return f"{days}d temu"
    if days <= 30:
        return f"{days // 7}tyg temu"
    return f"{days // 30}mies temu"


def _print_summary(name: str, total: int, recent: int, date_dist: Counter):
    pct = recent * 100 // max(total, 1)
    logger.info(f"[{name}] łącznie={total}  ostatnie {SCRAPER_DAYS}d={recent} ({pct}%)")
    top = date_dist.most_common(7)
    for label, cnt in top:
        bar = "█" * min(cnt, 30)
        logger.info(f"  {label:15s} {cnt:4d} {bar}")


# ── Testy per scraper ─────────────────────────────────────────────────────────

def test_jjit() -> list[dict]:
    from modules.scraper import _JJIT_API, JJIT_SEARCH_TERMS, _jjit_session

    session = _jjit_session()
    rows = []
    total = recent_n = 0

    for term in JJIT_SEARCH_TERMS[:2]:   # 2 terminy wystarczą do diagnozy
        cursor = None
        while True:
            try:
                url = (f"{_JJIT_API}?keywords={term.replace(' ', '+')}"
                       f"&keywordType=any&pageSize=20"
                       + (f"&from={cursor}" if cursor else ""))
                r = session.get(url, timeout=20)
                r.raise_for_status()
                raw = r.json()
            except Exception as e:
                logger.warning(f"[JJIT debug] {e}")
                break

            items = raw.get("data", [])
            meta  = raw.get("meta", {})
            total_int = meta.get("totalItems") or 0

            new_n = 0
            for item in items:
                pub = item.get("publishedAt")
                dt  = _parse_date(pub)
                passes = _is_recent(pub)
                total += 1
                if passes:
                    recent_n += 1
                rows.append({
                    "scraper":     "JJIT",
                    "term":        term,
                    "title":       item.get("title", ""),
                    "company":     item.get("companyName", ""),
                    "published_at": str(dt.date()) if dt else "",
                    "age":         _age_label(dt),
                    "passes":      "TAK" if passes else "NIE",
                })
                new_n += 1

            cursor = (meta.get("next") or {}).get("cursor")
            if new_n == 0 or cursor is None or (total_int and cursor >= total_int):
                break
            time.sleep(0.3)

    dist = Counter(r["age"] for r in rows)
    _print_summary("JJIT", total, recent_n, dist)
    return rows


def test_remoteok() -> list[dict]:
    rows = []
    total = recent_n = 0
    tags = ["product", "manager", "exec"]

    for tag in tags:
        try:
            r = requests.get(
                "https://remoteok.com/api",
                params={"tag": tag},
                headers=HEADERS,
                timeout=25,
            )
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            logger.warning(f"[RemoteOK debug] {e}")
            continue

        for job in data:
            if not isinstance(job, dict) or "id" not in job:
                continue
            date_val = job.get("epoch") or job.get("date")
            dt      = _parse_date(date_val)
            passes  = _is_recent(date_val)
            total  += 1
            if passes:
                recent_n += 1
            rows.append({
                "scraper":     "RemoteOK",
                "term":        tag,
                "title":       job.get("position", ""),
                "company":     job.get("company", ""),
                "published_at": str(dt.date()) if dt else "",
                "age":         _age_label(dt),
                "passes":      "TAK" if passes else "NIE",
            })
        time.sleep(1)

    dist = Counter(r["age"] for r in rows)
    _print_summary("RemoteOK", total, recent_n, dist)
    return rows


def test_wwr() -> list[dict]:
    rows = []
    total = recent_n = 0
    feeds = [
        "https://weworkremotely.com/categories/remote-management-and-finance-jobs.rss",
        "https://weworkremotely.com/categories/remote-product-jobs.rss",
    ]
    _WWR_NS = "https://weworkremotely.com/namespaces/rss/1.0"

    for feed_url in feeds:
        try:
            r = requests.get(feed_url, timeout=20)
            root = ET.fromstring(r.content)
        except Exception as e:
            logger.warning(f"[WWR debug] {e}")
            continue

        channel = root.find("channel")
        if not channel:
            continue

        for item in channel.findall("item"):
            pub    = item.findtext("pubDate")
            dt     = _parse_date(pub)
            passes = _is_recent(pub)
            total += 1
            if passes:
                recent_n += 1
            raw_title = item.findtext("title") or ""
            company, title = (raw_title.split(": ", 1) if ": " in raw_title
                              else ("", raw_title))
            rows.append({
                "scraper":     "WeWorkRemotely",
                "term":        feed_url.split("/")[-1].replace(".rss", ""),
                "title":       title.strip(),
                "company":     company.strip(),
                "published_at": str(dt.date()) if dt else "",
                "age":         _age_label(dt),
                "passes":      "TAK" if passes else "NIE",
            })

    dist = Counter(r["age"] for r in rows)
    _print_summary("WeWorkRemotely", total, recent_n, dist)
    return rows


def test_himalayas() -> list[dict]:
    rows = []
    total = recent_n = 0
    terms = ["product manager", "engineering manager"]

    for term in terms:
        try:
            r = requests.get(
                "https://himalayas.app/jobs/api",
                params={"q": term, "limit": 40, "offset": 0},
                headers=HEADERS,
                timeout=20,
            )
            r.raise_for_status()
            jobs = r.json().get("jobs", [])
        except Exception as e:
            logger.warning(f"[Himalayas debug] {e}")
            continue

        for job in jobs:
            date_val = job.get("pubDate")
            dt      = _parse_date(date_val)
            passes  = _is_recent(date_val)
            total  += 1
            if passes:
                recent_n += 1
            rows.append({
                "scraper":     "Himalayas",
                "term":        term,
                "title":       job.get("title", ""),
                "company":     job.get("companyName", ""),
                "published_at": str(dt.date()) if dt else "",
                "age":         _age_label(dt),
                "passes":      "TAK" if passes else "NIE",
            })
        time.sleep(0.8)

    dist = Counter(r["age"] for r in rows)
    _print_summary("Himalayas", total, recent_n, dist)
    return rows


# ── XLSX ──────────────────────────────────────────────────────────────────────

_COLS = [
    ("scraper",      14),
    ("term",         20),
    ("title",        35),
    ("company",      22),
    ("published_at", 12),
    ("age",          14),
    ("passes",       8),
]


def _save_xlsx(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Date filter debug"

    hdr_fill  = PatternFill("solid", fgColor="1A73E8")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    pass_fill = PatternFill("solid", fgColor="E8F5E9")
    fail_fill = PatternFill("solid", fgColor="FFF3E0")
    wrap   = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="top")

    for ci, (header, width) in enumerate(_COLS, 1):
        cell           = ws.cell(row=1, column=ci, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = "A2"

    for ri, row in enumerate(rows, 2):
        fill = pass_fill if row.get("passes") == "TAK" else fail_fill
        for ci, (key, _) in enumerate(_COLS, 1):
            cell           = ws.cell(row=ri, column=ci, value=row.get(key, ""))
            cell.fill      = fill
            cell.alignment = center if ci in {1, 5, 6, 7} else wrap
        ws.row_dimensions[ri].height = 18

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}1"
    wb.save(path)
    logger.info(f"[XLSX] Zapisano: {path}  ({len(rows)} wierszy)")


# ── Main ──────────────────────────────────────────────────────────────────────

def run() -> None:
    DEBUG_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    logger.info(f"=== DEBUG DATE FILTER (ostatnie {SCRAPER_DAYS} dni) ===")

    all_rows = []
    scrapers = [
        ("JJIT",           test_jjit),
        ("RemoteOK",       test_remoteok),
        ("WeWorkRemotely", test_wwr),
        ("Himalayas",      test_himalayas),
    ]

    for name, fn in scrapers:
        logger.info(f"--- {name} ---")
        rows = fn()
        all_rows.extend(rows)

    total   = len(all_rows)
    passing = sum(1 for r in all_rows if r["passes"] == "TAK")
    logger.info("=" * 60)
    logger.info(f"ŁĄCZNIE: {total} ofert  →  przechodzi filtr: {passing} ({passing*100//max(total,1)}%)")

    per_scraper = {}
    for r in all_rows:
        s = r["scraper"]
        per_scraper.setdefault(s, {"total": 0, "pass": 0})
        per_scraper[s]["total"] += 1
        if r["passes"] == "TAK":
            per_scraper[s]["pass"] += 1
    for s, d in per_scraper.items():
        pct = d["pass"] * 100 // max(d["total"], 1)
        logger.info(f"  {s:18s} {d['pass']:4d}/{d['total']:4d}  ({pct}%)")
    logger.info("=" * 60)

    path = DEBUG_DIR / f"date_filter_{ts}.xlsx"
    _save_xlsx(all_rows, path)
    logger.info(f"XLSX → {path.resolve()}")


if __name__ == "__main__":
    run()
