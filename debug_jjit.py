"""
Moduł debugowy scrapera JustJoinIT.
Wywołuje scrape_justjoinit() i zapisuje wyniki do XLSX w ./debug/.

Uruchomienie:
    py -3.12 debug_jjit.py
    py -3.12 debug_jjit.py --limit 20   # ogranicz liczbę opisów

NIE modyfikuje main.py ani bazy danych.
"""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Force UTF-8 na Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

DEBUG_DIR = Path("debug")

# ── Kolumny XLSX ──────────────────────────────────────────────────────────────

_COLS = [
    ("title",           32),
    ("company",         22),
    ("location",        20),
    ("salary",          30),
    ("salary_from",     12),
    ("salary_contract", 14),
    ("skills",          40),
    ("url",             55),
    ("desc_len",        10),
    ("description",     80),
]
_CENTER = {3, 5, 6, 9}  # kolumny wyrównane do środka (1-indexed)


def _save_xlsx(offers: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JJIT scraper"

    hdr_fill = PatternFill("solid", fgColor="1A73E8")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    wrap     = Alignment(vertical="top", wrap_text=True)
    center   = Alignment(horizontal="center", vertical="top")

    for ci, (header, width) in enumerate(_COLS, 1):
        cell           = ws.cell(row=1, column=ci, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = "A2"

    url_font = Font(color="1A73E8", underline="single", size=9)

    for ri, o in enumerate(offers, 2):
        desc = o.get("description") or ""
        row_data = [
            o.get("title", ""),
            o.get("company", ""),
            o.get("location", ""),
            o.get("salary", ""),
            o.get("salary_from") or 0,
            o.get("salary_contract", ""),
            ", ".join(o.get("skills") or []),
            o.get("url", ""),
            len(desc),
            desc[:500],
        ]
        for ci, val in enumerate(row_data, 1):
            cell           = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = center if ci in _CENTER else wrap

        # URL klikalny
        url_val = o.get("url", "")
        if url_val:
            c          = ws.cell(row=ri, column=8)
            c.hyperlink = url_val
            c.font      = url_font

        ws.row_dimensions[ri].height = 40 if desc else 20

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}1"
    wb.save(path)
    logger.info(f"[XLSX] Zapisano: {path}  ({len(offers)} wierszy)")


# ── Główna funkcja ────────────────────────────────────────────────────────────

def run(limit: int | None = None) -> None:
    from modules.scraper import scrape_justjoinit

    DEBUG_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    logger.info("=== DEBUG JJIT SCRAPER START ===")
    offers = scrape_justjoinit()

    if limit:
        offers = offers[:limit]
        logger.info(f"Ograniczono do {limit} ofert (--limit)")

    # ── Statystyki ────────────────────────────────────────────────────────────
    n_total   = len(offers)
    n_salary  = sum(1 for o in offers if o.get("salary_from", 0) > 0)
    n_desc    = sum(1 for o in offers if o.get("description"))
    n_remote  = sum(1 for o in offers if "remote" in (o.get("location") or "").lower())
    n_hybrid  = sum(1 for o in offers if "hybrid" in (o.get("location") or "").lower())

    avg_desc  = (
        sum(len(o.get("description") or "") for o in offers if o.get("description"))
        // max(n_desc, 1)
    )

    logger.info("=" * 60)
    logger.info(f"WYNIKI  –  {n_total} ofert łącznie")
    logger.info(f"  z salary_from > 0 : {n_salary}  ({n_salary*100//max(n_total,1)}%)")
    logger.info(f"  z opisem          : {n_desc}  ({n_desc*100//max(n_total,1)}%)")
    logger.info(f"  avg opis [zn.]    : {avg_desc}")
    logger.info(f"  remote            : {n_remote}")
    logger.info(f"  hybrid            : {n_hybrid}")
    logger.info("=" * 60)

    # ── Podgląd w terminalu ───────────────────────────────────────────────────
    print(f"\n{'#':>3}  {'Score':>5}  {'Tytuł':<40} {'Firma':<22} {'Lokacja':<22} {'Salary'}")
    print("-" * 130)
    for i, o in enumerate(offers, 1):
        title    = (o.get("title") or "")[:38]
        company  = (o.get("company") or "")[:20]
        loc      = (o.get("location") or "")[:20]
        sal      = (o.get("salary") or "—")[:30]
        desc_ind = "📄" if o.get("description") else "  "
        print(f"{i:>3}  {desc_ind}    {title:<40} {company:<22} {loc:<22} {sal}")

    # ── XLSX ──────────────────────────────────────────────────────────────────
    xlsx_path = DEBUG_DIR / f"jjit_scraper_{ts}.xlsx"
    _save_xlsx(offers, xlsx_path)

    logger.info(f"XLSX → {xlsx_path.resolve()}")
    logger.info("=== DEBUG JJIT SCRAPER KONIEC ===")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Debug scraper JustJoinIT")
    parser.add_argument("--limit", type=int, default=None,
                        help="Ogranicz liczbę wynikowych ofert (nie ogranicza scrapingu)")
    args = parser.parse_args()
    run(limit=args.limit)
